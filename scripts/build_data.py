import json
import re
import socket
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from google_play_scraper import Sort, app as gp_app, reviews
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

socket.setdefaulttimeout(25)

BASE = Path(__file__).resolve().parents[1]
DATA_DIR = BASE / "data"
JSON_PATH = DATA_DIR / "dashboard_data.json"
WORKBOOK_PATH = BASE / "healthify_reputation_data.xlsx"

PLAY_ID = "com.healthifyme.basic"
APPLE_ID = "943712366"
TRUSTPILOT_URLS = [
    "https://www.trustpilot.com/review/healthifyme.com",
    "https://www.trustpilot.com/review/www.healthifyme.com",
]
COUNTRIES = {
    "in": {"label": "India", "bucket": "India"},
    "us": {"label": "United States", "bucket": "Outside India"},
    "gb": {"label": "United Kingdom", "bucket": "Outside India"},
    "au": {"label": "Australia", "bucket": "Outside India"},
    "ae": {"label": "United Arab Emirates", "bucket": "Outside India"},
    "sg": {"label": "Singapore", "bucket": "Outside India"},
}
PLAY_REVIEW_CAPS = {
    "in": 10000,
    "us": 1000,
    "gb": 1000,
    "au": 1000,
    "ae": 1000,
    "sg": 1000,
}
MAX_APPSTORE_RSS_PAGES = 10

THEMES = {
    "Food Logging": ["food", "calorie", "calories", "meal", "log", "logging", "nutrition", "macro", "protein"],
    "Coach / Support": ["coach", "coaching", "trainer", "dietician", "nutritionist", "support", "expert"],
    "Weight Loss": ["weight", "loss", "lose", "lost", "journey", "progress", "goal"],
    "Subscription / Payment": ["subscription", "payment", "refund", "paid", "charge", "cancel", "price", "money", "premium"],
    "Bugs / Reliability": ["bug", "crash", "error", "slow", "login", "stuck", "broken", "issue", "problem"],
    "UX": ["easy", "interface", "ui", "simple", "confusing", "design", "feature"],
    "Data Accuracy": ["accurate", "accuracy", "database", "wrong", "scan", "barcode", "steps", "sync"],
    "Customer Service": ["customer", "service", "response", "help", "called", "contact"],
}


def clean_text(value):
    value = (value or "").strip()
    try:
        value = value.encode("latin1").decode("utf-8")
    except Exception:
        pass
    return " ".join(value.split())


def theme_for(text):
    lower = (text or "").lower()
    matches = []
    for theme, words in THEMES.items():
        if any(word in lower for word in words):
            matches.append(theme)
    return matches[0] if matches else "General"


def sentiment_for(rating):
    try:
        value = int(float(rating))
    except Exception:
        return "Unknown"
    if value >= 4:
        return "Positive"
    if value <= 2:
        return "Negative"
    return "Mixed"


def fetch_play_meta(country):
    result = gp_app(PLAY_ID, lang="en", country=country)
    return {
        "source": "Google Play",
        "country": country,
        "region_label": COUNTRIES[country]["label"],
        "region_bucket": COUNTRIES[country]["bucket"],
        "rating": result.get("score"),
        "rating_count": result.get("ratings") or result.get("reviews"),
        "review_count": result.get("reviews"),
        "install_band": result.get("installs"),
        "note": "Google Play exposes install bands, not exact installs.",
        "icon": result.get("icon"),
        "screenshots": (result.get("screenshots") or [])[:8],
        "url": f"https://play.google.com/store/apps/details?id={PLAY_ID}&hl=en&gl={country.upper()}",
    }


def fetch_play_reviews(country):
    rows = []
    token = None
    seen = set()
    zero_batches = 0
    failed_batches = 0
    cap = PLAY_REVIEW_CAPS.get(country, 1000)
    while len(rows) < cap:
        try:
            batch, token = reviews(
                PLAY_ID,
                lang="en",
                country=country,
                sort=Sort.NEWEST,
                count=min(200, cap - len(rows)),
                continuation_token=token,
            )
        except Exception as exc:
            failed_batches += 1
            print(f"  {country}: review batch failed ({failed_batches}/3): {exc}", flush=True)
            if failed_batches >= 3:
                break
            time.sleep(2)
            continue
        failed_batches = 0
        if not batch:
            zero_batches += 1
            if zero_batches >= 2 or not token:
                break
            continue
        zero_batches = 0
        new_count = 0
        for item in batch:
            rid = item.get("reviewId")
            if rid in seen:
                continue
            seen.add(rid)
            content = clean_text(item.get("content"))
            rows.append({
                "source": "Google Play",
                "country": country,
                "region_label": COUNTRIES[country]["label"],
                "region_bucket": COUNTRIES[country]["bucket"],
                "review_id": rid,
                "author": item.get("userName"),
                "title": None,
                "rating": item.get("score"),
                "date": item.get("at").isoformat() if item.get("at") else None,
                "content": content,
                "theme_primary": theme_for(content),
                "sentiment": sentiment_for(item.get("score")),
                "version": item.get("appVersion") or item.get("reviewCreatedVersion"),
            })
            new_count += 1
            if len(rows) >= cap:
                break
        if new_count == 0 or not token:
            break
        print(f"  {country}: {len(rows)}/{cap} Play reviews", flush=True)
    return rows[:cap]


def fetch_apple_meta(country):
    result = requests.get("https://itunes.apple.com/lookup", params={"id": APPLE_ID, "country": country}, timeout=30).json().get("results", [{}])[0]
    return {
        "source": "App Store",
        "country": country,
        "region_label": COUNTRIES[country]["label"],
        "region_bucket": COUNTRIES[country]["bucket"],
        "rating": result.get("averageUserRating"),
        "rating_count": result.get("userRatingCount"),
        "review_count": None,
        "install_band": None,
        "note": "Apple does not expose public download counts.",
        "icon": result.get("artworkUrl512") or result.get("artworkUrl100"),
        "screenshots": result.get("screenshotUrls") or [],
        "url": result.get("trackViewUrl") or f"https://apps.apple.com/app/id{APPLE_ID}",
    }


def fetch_apple_rss_reviews(country):
    rows = []
    if MAX_APPSTORE_RSS_PAGES <= 0:
        return rows
    for page in range(1, MAX_APPSTORE_RSS_PAGES + 1):
        url = f"https://itunes.apple.com/{country}/rss/customerreviews/page={page}/id={APPLE_ID}/sortby=mostrecent/json"
        try:
            data = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"}).json()
        except Exception as exc:
            print(f"  {country}: App Store RSS page {page} failed: {exc}", flush=True)
            break
        entries = data.get("feed", {}).get("entry", [])
        if not entries:
            break
        added = 0
        for entry in entries:
            if "content" not in entry:
                continue
            content = clean_text(entry.get("content", {}).get("label"))
            rating = entry.get("im:rating", {}).get("label")
            rows.append({
                "source": "App Store",
                "country": country,
                "region_label": COUNTRIES[country]["label"],
                "region_bucket": COUNTRIES[country]["bucket"],
                "review_id": entry.get("id", {}).get("label"),
                "author": entry.get("author", {}).get("name", {}).get("label"),
                "title": clean_text(entry.get("title", {}).get("label")),
                "rating": int(rating) if rating and rating.isdigit() else None,
                "date": entry.get("updated", {}).get("label"),
                "content": content,
                "theme_primary": theme_for(content),
                "sentiment": sentiment_for(rating),
                "version": entry.get("im:version", {}).get("label"),
            })
            added += 1
        if added == 0:
            break
        print(f"  {country}: {len(rows)} App Store RSS reviews", flush=True)
    return rows


def fetch_apple_visible_reviews(country):
    url = f"https://apps.apple.com/{country}/app/id{APPLE_ID}?see-all=reviews"
    html = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"}).text
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    for card in soup.select('li > div.container[aria-labelledby]'):
        title = card.select_one("h3.title")
        author = card.select_one("p.author")
        body = card.select_one('div.content p[data-testid="truncate-text"]')
        stars = card.select_one("ol.stars[aria-label]")
        date = card.select_one("time.date")
        if not (title and author and body):
            continue
        label = stars.get("aria-label") if stars else ""
        rating = int(label[0]) if label and label[0].isdigit() else None
        content = clean_text(body.get_text(" ", strip=True))
        review_id = card.get("aria-labelledby", "").replace("review-", "").replace("-title", "") or None
        rows.append({
            "source": "App Store",
            "country": country,
            "region_label": COUNTRIES[country]["label"],
            "region_bucket": COUNTRIES[country]["bucket"],
            "review_id": review_id,
            "author": clean_text(author.get_text(" ", strip=True)),
            "title": clean_text(title.get_text(" ", strip=True)),
            "rating": rating,
            "date": date.get("datetime") if date else None,
            "content": content,
            "theme_primary": theme_for(content),
            "sentiment": sentiment_for(rating),
            "version": None,
        })
    return rows


def fetch_apple_reviews(country):
    rss = fetch_apple_rss_reviews(country)
    visible = fetch_apple_visible_reviews(country)
    by_id = {}
    for row in rss + visible:
        key = row.get("review_id") or f"{row.get('author')}|{row.get('title')}|{row.get('date')}"
        by_id[key] = row
    return list(by_id.values())


def dedupe_reviews(rows):
    seen = set()
    deduped = []
    for row in rows:
        key = (
            row.get("source"),
            row.get("review_id")
            or f"{row.get('country')}|{row.get('author')}|{row.get('title')}|{row.get('date')}|{row.get('content')}"
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def fetch_trustpilot():
    sources = []
    reviews_out = []
    for url in TRUSTPILOT_URLS:
        try:
            response = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        except Exception as exc:
            sources.append({"source": "Trustpilot", "region_label": "Global", "region_bucket": "All", "rating": None, "rating_count": None, "review_count": 0, "install_band": None, "note": f"Trustpilot request failed: {exc}", "url": url})
            continue
        if response.status_code != 200:
            sources.append({"source": "Trustpilot", "region_label": "Global", "region_bucket": "All", "rating": None, "rating_count": None, "review_count": 0, "install_band": None, "note": f"Trustpilot returned HTTP {response.status_code}; direct public scraping was unavailable from this environment.", "url": url})
            continue
        soup = BeautifulSoup(response.text, "html.parser")
        json_ld = soup.find("script", type="application/ld+json")
        sources.append({"source": "Trustpilot", "region_label": "Global", "region_bucket": "All", "rating": None, "rating_count": None, "review_count": len(reviews_out), "install_band": None, "note": "Trustpilot page loaded, but structured review parsing is best effort.", "url": url})
        break
    if not sources:
        sources.append({"source": "Trustpilot", "region_label": "Global", "region_bucket": "All", "rating": None, "rating_count": None, "review_count": 0, "install_band": None, "note": "No Trustpilot profile resolved.", "url": TRUSTPILOT_URLS[0]})
    return sources, reviews_out


def write_workbook(sources, reviews_out):
    wb = Workbook()
    header_fill = "1F4E78"
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    ws.append(["Generated at UTC", datetime.now(timezone.utc).isoformat()])
    ws.append(["Play Store package", PLAY_ID])
    ws.append(["App Store ID", APPLE_ID])
    ws.append(["Countries sampled", ", ".join(COUNTRIES.keys())])
    ws.append(["Trustpilot note", next((s.get("note") for s in sources if s.get("source") == "Trustpilot"), "")])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = header_font
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110

    for title, rows in [("Sources", sources), ("Reviews", reviews_out)]:
        sheet = wb.create_sheet(title)
        headers = sorted({key for row in rows for key in row.keys()}) if rows else ["note"]
        sheet.append(headers)
        for row in rows:
            values = []
            for header in headers:
                value = row.get(header)
                if isinstance(value, (list, dict)):
                    value = json.dumps(value)
                values.append(value)
            sheet.append(values)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=header_fill)
            cell.font = header_font
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 22 if col[0].value != "content" else 120
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    wb.save(WORKBOOK_PATH)


def build():
    DATA_DIR.mkdir(exist_ok=True)
    sources = []
    reviews_out = []

    for country in COUNTRIES:
        print(f"Fetching Google Play metadata/reviews for {country}...")
        play_meta = fetch_play_meta(country)
        play_reviews = fetch_play_reviews(country)
        play_meta["review_count"] = len(play_reviews)
        sources.append(play_meta)
        reviews_out.extend(play_reviews)

        print(f"Fetching App Store metadata/reviews for {country}...")
        apple_meta = fetch_apple_meta(country)
        apple_reviews = fetch_apple_reviews(country)
        apple_meta["review_count"] = len(apple_reviews)
        sources.append(apple_meta)
        reviews_out.extend(apple_reviews)

    trust_sources, trust_reviews = fetch_trustpilot()
    sources.extend(trust_sources)
    reviews_out.extend(trust_reviews)
    reviews_out = dedupe_reviews(reviews_out)

    rating_counts = Counter(str(int(r["rating"])) for r in reviews_out if r.get("rating"))
    theme_counts = Counter(r.get("theme_primary") or "General" for r in reviews_out)
    visuals = []
    for source in sources:
        if source.get("source") in {"Google Play", "App Store"} and (source.get("icon") or source.get("screenshots")):
            visuals.append({"source": source["source"], "region_label": source["region_label"], "icon": source.get("icon"), "screenshots": source.get("screenshots") or []})

    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "sources": sources,
        "reviews": reviews_out,
        "visuals": visuals[:8],
        "summary": {
            "review_count": len(reviews_out),
            "rating_breakdown": dict(rating_counts),
            "top_themes": [{"theme": k, "count": v} for k, v in theme_counts.most_common(12)],
        },
        "notes": [
            "India vs outside India uses sampled public storefronts: IN, US, GB, AU, AE, SG.",
            "Google Play exposes install bands, not exact downloads.",
            "Apple does not expose public downloads.",
            "Apple App Store reviews use the public RSS customer review feed where available; Apple does not expose full private review history or downloads.",
            "Trustpilot blocked direct public scraping from this environment during this build; the dashboard records that source availability explicitly.",
        ],
    }
    JSON_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    write_workbook(sources, reviews_out)
    print(f"Wrote {JSON_PATH}")
    print(f"Wrote {WORKBOOK_PATH}")
    print(f"Reviews: {len(reviews_out)}")


if __name__ == "__main__":
    build()
